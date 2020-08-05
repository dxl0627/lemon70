

import openpyxl
import requests



def read_data(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row
    # print(max_row)
    case_list = []
    for i in range(2, max_row + 1, 1):
        dict1 = dict(
            case_id=sheet.cell(row=i, column=1).value,
            url=sheet.cell(row=i, column=5).value,
            data=sheet.cell(row=i, column=6).value,
            expected=sheet.cell(row=i, column=7).value
        )
        case_list.append(dict1)
    return case_list



def api_func(url, data):
    header_login = {'X-Lemonban-Media-Type': 'lemonban.v2',
                    'Content-Type': 'application/json'}
    res1 = requests.post(url=url, json=data, headers=header_login)
    # print(res1.json())
    response = res1.json()
    return response



def write_result(filename, sheetname, row, column, final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = final_result
    wb.save(filename)



def execute_func(filename, sheetname):
    cases = read_data(filename, sheetname)
    # print(cases)
    for case in cases:
        case_id = case.get('case_id')
        url = case['url']
        data = case.get('data')
        data = eval(data)
        expected = case.get('expected')
        expected = eval(expected)
        expected_msg = expected.get('msg')
        real_result = api_func(url=url, data=data)
        real_msg = real_result.get('msg')
        print('Ԥ�ڽ��Ϊ��{}'.format(expected_msg))
        print('ʵ�ʽ��Ϊ��{}'.format(real_msg))
        if real_msg == expected_msg:
            print('��{}������ͨ����'.format(case_id))
            final_res = 'pass'
        else:
            print('��{}������δͨ����'.format(case_id))
            final_res = 'fail'
        print('*' * 30)
        write_result(filename, sheetname, case_id + 1, 8, final_res)


execute_func('test_case_api.xlsx', 'register')
execute_func('test_case_api.xlsx', 'login')
